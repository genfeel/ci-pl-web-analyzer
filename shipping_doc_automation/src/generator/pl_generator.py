"""PL(Packing List) 엑셀 생성기

샘플과 동일한 레이아웃으로 PL 엑셀을 생성.
CI 원본 파일에서 헤더/서명 정보를 복사하고, 케이스 데이터를 채움.
"""
import re
from pathlib import Path
from typing import List, Optional
from copy import copy

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.data_models import PackedCase, PLDocument, ProductCategory
from src.generator.format_rules import (
    get_category_display_name,
    format_dimensions,
    format_cbm,
    format_weight,
    format_case_header,
    format_total_header,
)
from config.packing_constraints import CATEGORY_PACKING_PRIORITY


# 스타일 정의
BOLD_FONT = Font(name='Arial', size=10, bold=True)
NORMAL_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
HEADER_FONT = Font(name='Arial', size=9, bold=True)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def generate_pl_excel(
    pl_doc: PLDocument,
    ci_filepath: str,
    output_path: str,
    template_filepath: Optional[str] = None,
) -> str:
    """PL 엑셀 파일 생성

    Args:
        pl_doc: 생성된 PL 문서 객체
        ci_filepath: CI 원본 파일 경로 (헤더 정보 복사용)
        output_path: 출력 파일 경로
        template_filepath: 템플릿 파일 경로 (없으면 새로 생성)

    Returns:
        생성된 파일 경로
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PL"

    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 22

    total_cases = pl_doc.total_cases

    # 헤더 영역 작성
    row = _write_header(ws, pl_doc, ci_filepath)

    # 마킹 정보 영역
    row = _write_marking_info(ws, row, total_cases)

    # 케이스별 데이터 작성
    sheet_num = 1
    cases_on_page = 0
    max_rows_per_page = 45  # 페이지당 최대 행 (대략)
    page_start_row = row

    for case in pl_doc.cases:
        # 페이지 넘김 체크
        estimated_rows = 2 + len(case.items)  # 케이스 헤더 + 카테고리 + 품목
        if cases_on_page > 0 and (row - page_start_row + estimated_rows) > max_rows_per_page:
            # "- to be continued -" 추가
            row += 1
            ws.cell(row=row, column=4, value="- to be continued -").font = BOLD_FONT
            row += 2

            # 서명 영역
            row = _write_signature(ws, row)
            row += 2

            # 새 페이지 헤더
            sheet_num += 1
            row = _write_page_header(ws, row, pl_doc, sheet_num)
            row = _write_column_headers(ws, row)
            page_start_row = row
            cases_on_page = 0

        # 케이스 데이터 작성
        row = _write_case(ws, row, case, total_cases, pl_doc.is_combined_order)
        cases_on_page += 1

    # TOTAL 행
    row += 1
    row = _write_total(ws, row, pl_doc)

    # 서명 영역
    row += 2
    _write_signature(ws, row)

    # 저장
    output_path = str(output_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return output_path


def _write_header(ws, pl_doc: PLDocument, ci_filepath: str) -> int:
    """PL 헤더 영역 작성 (Row 1~28)"""
    row = 1

    # 타이틀
    ws.cell(row=row, column=1, value="P A C K I N G    L I S T").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    row += 1

    # Sheet 정보
    total_sheets = max(1, len(pl_doc.cases) // 8 + 1)
    ws.cell(row=row, column=8, value=f"Sheet 1 of {total_sheets}").font = BOLD_FONT
    ws.cell(row=row, column=8).alignment = RIGHT_ALIGN
    row += 1

    # 송하인 정보
    ws.cell(row=row, column=1, value="Shipper / Exporter").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="HD HYUNDAI ELECTRIC CO., LTD.").font = BOLD_FONT
    row += 1
    ws.cell(row=row, column=1, value="16, Seongseo-ro, Dalseo-gu,").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Daegu, 42704, KOREA").font = NORMAL_FONT
    row += 2

    # Invoice 정보
    inv_no = pl_doc.header_info.get('invoice_no', '')
    if not inv_no and pl_doc.order_numbers:
        inv_no = pl_doc.order_numbers[0] if not pl_doc.is_combined_order else ' / '.join(pl_doc.order_numbers)
    ws.cell(row=5, column=5, value="Invoice No.").font = HEADER_FONT
    ws.cell(row=5, column=8, value=inv_no).font = NORMAL_FONT

    # 수하인
    ws.cell(row=row, column=1, value="For Account & Risk of Messrs.").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value=pl_doc.header_info.get('consignee', '')).font = NORMAL_FONT
    row += 5

    # Notify party
    ws.cell(row=row, column=1, value="Notify party").font = HEADER_FONT
    row += 6

    # Port
    ws.cell(row=row, column=1, value="Port of loading").font = HEADER_FONT
    ws.cell(row=row, column=5, value="Final destination").font = HEADER_FONT
    row += 2

    # Carrier
    ws.cell(row=row, column=1, value="Per").font = HEADER_FONT
    ws.cell(row=row, column=5, value="Sailing on or about").font = HEADER_FONT
    row += 2

    # 컬럼 헤더
    row = _write_column_headers(ws, row)

    return row


def _write_column_headers(ws, row: int) -> int:
    """PL 컬럼 헤더 작성"""
    headers = [
        (1, "Marks and\nNo. of\npackages"),
        (3, "Description of Goods"),
        (5, "Quantity"),
        (6, "Net-weight\n(kg)"),
        (7, "Gross-weight\n(kg)"),
        (8, "Measurement\n(LXWXH.MM)"),
    ]

    for col, text in headers:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    row += 1
    return row


def _write_marking_info(ws, row: int, total_cases: int) -> int:
    """마킹 정보 작성"""
    marking_lines = [
        f"CASE NO. 1~{total_cases}",
        "GROSS WEIGHT:",
        "DIMENSIONS:",
        "",
        "MADE IN SOUTH KOREA",
    ]

    for line in marking_lines:
        ws.cell(row=row, column=1, value=line).font = NORMAL_FONT
        row += 1

    return row


def _write_case(
    ws, row: int, case: PackedCase, total_cases: int, is_combined: bool
) -> int:
    """개별 케이스 데이터 작성"""
    # 빈 행
    row += 1

    # 케이스 헤더 행
    case_header = format_case_header(case.case_no, total_cases)

    if is_combined and case.order_number:
        # 합적: A열에 주문번호
        ws.cell(row=row, column=1, value=case.order_number).font = BOLD_FONT
        ws.cell(row=row, column=3, value=case_header).font = BOLD_FONT
        ws.cell(row=row, column=5, value="(S-TOTAL)").font = BOLD_FONT
    else:
        ws.cell(row=row, column=3, value=case_header).font = BOLD_FONT

    # 케이스 헤더에 중량/규격
    ws.cell(row=row, column=6, value=case.net_weight).font = BOLD_FONT
    ws.cell(row=row, column=6).number_format = '0.00'
    ws.cell(row=row, column=7, value=case.gross_weight).font = BOLD_FONT
    ws.cell(row=row, column=7).number_format = '0'
    ws.cell(row=row, column=8, value=format_dimensions(case.dimensions)).font = BOLD_FONT

    row += 1

    # 카테고리별 품목 그룹핑
    cat_groups = {}
    for pi in case.items:
        cat = pi.ci_item.category
        if cat not in cat_groups:
            cat_groups[cat] = []
        cat_groups[cat].append(pi)

    first_category = True
    for cat in sorted(cat_groups.keys(), key=lambda c: CATEGORY_PACKING_PRIORITY.get(c.value, 99)):
        items = cat_groups[cat]

        # 카테고리 헤더
        cat_name = get_category_display_name(cat, plural=True)
        ws.cell(row=row, column=3, value=cat_name).font = BOLD_FONT

        # 첫 카테고리 행에 CBM
        if first_category:
            cbm = format_cbm(case.cbm)
            if cbm:
                ws.cell(row=row, column=8, value=float(cbm)).font = BOLD_FONT
                ws.cell(row=row, column=8).number_format = '0.000'
            first_category = False

        row += 1

        # 품목 행
        for pi in items:
            ws.cell(row=row, column=3, value=pi.ci_item.description).font = NORMAL_FONT
            ws.cell(row=row, column=5, value=pi.quantity).font = NORMAL_FONT

            # 합적 주문: 개별 품목 순중량 표시
            if is_combined and pi.net_weight > 0:
                ws.cell(row=row, column=6, value=pi.net_weight).font = NORMAL_FONT
                ws.cell(row=row, column=6).number_format = '0.00'

            row += 1

    return row


def _write_total(ws, row: int, pl_doc: PLDocument) -> int:
    """TOTAL 행 작성"""
    total_header = format_total_header(pl_doc.total_cases)

    ws.cell(row=row, column=3, value=total_header).font = BOLD_FONT
    ws.cell(row=row, column=5, value=pl_doc.total_quantity).font = BOLD_FONT
    ws.cell(row=row, column=6, value=pl_doc.total_net_weight).font = BOLD_FONT
    ws.cell(row=row, column=6).number_format = '0.00'
    ws.cell(row=row, column=7, value=pl_doc.total_gross_weight).font = BOLD_FONT
    ws.cell(row=row, column=7).number_format = '0'
    ws.cell(row=row, column=8, value=pl_doc.total_cbm).font = BOLD_FONT
    ws.cell(row=row, column=8).number_format = '0.000'

    # 구분선
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = Border(top=Side(style='double'))

    return row + 1


def _write_page_header(ws, row: int, pl_doc: PLDocument, sheet_num: int) -> int:
    """페이지 헤더 (2페이지 이후)"""
    ws.cell(row=row, column=1, value="P A C K I N G    L I S T").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    row += 1

    total_sheets = max(1, len(pl_doc.cases) // 8 + 1)
    ws.cell(row=row, column=8, value=f"Sheet {sheet_num} of {total_sheets}").font = BOLD_FONT
    ws.cell(row=row, column=8).alignment = RIGHT_ALIGN
    row += 2

    return row


def _write_signature(ws, row: int) -> int:
    """서명 영역 작성"""
    ws.cell(row=row, column=1, value="HD HYUNDAI ELECTRIC CO., LTD.").font = BOLD_FONT
    row += 3
    ws.cell(row=row, column=1, value="___________________________").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Authorized Signature").font = NORMAL_FONT
    return row + 1


def generate_pl_from_template(
    pl_doc: PLDocument,
    ci_filepath: str,
    output_path: str,
) -> str:
    """CI 원본을 템플릿으로 활용하여 PL 생성

    CI 파일의 PL 시트를 복사하고 데이터만 교체하는 방식.
    더 정확한 서식 재현이 가능.
    """
    ext = Path(ci_filepath).suffix.lower()

    if ext == '.xlsx':
        return _generate_from_xlsx_template(pl_doc, ci_filepath, output_path)
    else:
        # .xls는 직접 생성
        return generate_pl_excel(pl_doc, ci_filepath, output_path)


def _generate_from_xlsx_template(
    pl_doc: PLDocument, ci_filepath: str, output_path: str
) -> str:
    """xlsx 파일의 PL 시트를 템플릿으로 활용"""
    wb = openpyxl.load_workbook(ci_filepath)

    # PL 시트 찾기
    pl_sheet = None
    for name in wb.sheetnames:
        if 'PL' in name.upper() and 'CI' not in name.upper():
            pl_sheet = name
            break
    if not pl_sheet:
        pl_sheet = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    ws = wb[pl_sheet]
    total_cases = pl_doc.total_cases

    # 기존 데이터 영역 찾기 (CASE NO 시작 ~ TOTAL)
    data_start = None
    data_end = None
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(row=r, column=3).value or '').upper()
        if 'CASE NO' in val and data_start is None:
            data_start = r
        if val.startswith('TOTAL') and ('CASES' in val or 'CASE' in val):
            data_end = r
            break

    if data_start is None:
        wb.close()
        return generate_pl_excel(pl_doc, ci_filepath, output_path)

    # 기존 데이터 삭제 (data_start ~ data_end)
    if data_end:
        for r in range(data_start, data_end + 10):
            for c in range(1, 9):
                ws.cell(row=r, column=c).value = None

    # 새 데이터 작성
    row = data_start
    for case in pl_doc.cases:
        row += 1
        row = _write_case(ws, row - 1, case, total_cases, pl_doc.is_combined_order)

    # TOTAL
    row += 1
    _write_total(ws, row, pl_doc)

    # 서명
    row += 2
    _write_signature(ws, row)

    # 저장
    output_path = str(output_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return output_path
