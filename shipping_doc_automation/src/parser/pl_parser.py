"""PL(Packing List) 엑셀 파서 - 검증용

샘플 PL을 파싱하여 자동생성된 PL과 비교하는 데 사용.
PL 구조:
- A=marks, C=description, E=quantity, F=net_weight, G=gross_weight, H=measurement
- 케이스 헤더: "CASE NO : X of Y" (C열)
- 카테고리: BOLD 텍스트 (C열)
- 합적: A열에 주문번호, E열에 (S-TOTAL)
"""
import re
from pathlib import Path
from typing import Optional

import openpyxl
import xlrd

from src.models.data_models import PackedCase, PackedItem, PLDocument, CILineItem, ProductCategory
from src.parser.model_number_parser import (
    extract_model_from_description,
    is_category_header,
    detect_category_from_header,
)


def _find_pl_sheet(sheet_names: list) -> Optional[str]:
    for name in sheet_names:
        upper = name.upper().strip()
        if upper == 'PL' or upper == 'P/L':
            return name
        if 'PL' in upper and 'CI' not in upper:
            return name
    for name in sheet_names:
        if 'PL' in name.upper() or 'P/L' in name.upper():
            return name
    return sheet_names[1] if len(sheet_names) > 1 else None


def _read_sheet(filepath: str, sheet_name: str) -> list:
    ext = Path(filepath).suffix.lower()
    if ext == '.xls':
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for r in range(ws.nrows):
            row_data = []
            ncols = min(ws.ncols, 10)
            for c in range(ncols):
                cell = ws.cell(r, c)
                xf_index = ws.cell_xf_index(r, c)
                xf = wb.xf_list[xf_index]
                font = wb.font_list[xf.font_index]
                row_data.append({
                    'value': cell.value if cell.value != '' else None,
                    'bold': font.bold,
                })
            while len(row_data) < 10:
                row_data.append({'value': None, 'bold': False})
            rows.append(row_data)
        wb.release_resources()
        return rows
    else:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10, values_only=False):
            row_data = []
            for cell in row:
                row_data.append({
                    'value': cell.value,
                    'bold': cell.font.bold if cell.font else False,
                })
            while len(row_data) < 10:
                row_data.append({'value': None, 'bold': False})
            rows.append(row_data)
        wb.close()
        return rows


def _cell_str(cell) -> str:
    val = cell.get('value')
    return str(val).strip() if val is not None else ''


def _cell_num(cell) -> Optional[float]:
    val = cell.get('value')
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


def _parse_case_header(text: str) -> Optional[tuple]:
    """CASE NO : X of Y → (X, Y)"""
    match = re.search(r'CASE\s*NO\s*[:.]\s*(\d+)\s*(?:of|/)\s*(\d+)', text, re.IGNORECASE)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def _parse_dimensions(text: str) -> Optional[tuple]:
    """1100X1100X1120.MM → (1100, 1100, 1120)"""
    match = re.search(r'(\d+)\s*[Xx×]\s*(\d+)\s*[Xx×]\s*(\d+)', str(text))
    if match:
        return (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def parse_pl(filepath: str) -> PLDocument:
    """PL 엑셀 파일 파싱 → PLDocument 반환"""
    filepath = str(filepath)
    ext = Path(filepath).suffix.lower()

    if ext == '.xls':
        wb = xlrd.open_workbook(filepath)
        sheet_names = wb.sheet_names()
        wb.release_resources()
    else:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

    pl_sheet = _find_pl_sheet(sheet_names)
    if not pl_sheet:
        raise ValueError(f"PL 시트를 찾을 수 없습니다: {sheet_names}")

    rows = _read_sheet(filepath, pl_sheet)

    doc = PLDocument(filename=Path(filepath).name)
    cases = []
    current_case = None
    current_category = ''
    skip_until_data = True

    for i, row in enumerate(rows):
        if len(row) < 8:
            continue

        a_val = _cell_str(row[0])
        c_val = _cell_str(row[2])
        e_val = _cell_num(row[4])
        f_val = _cell_num(row[5])
        g_val = _cell_num(row[6])
        h_val = _cell_str(row[7])

        # TOTAL 행 → 파싱 종료
        c_upper = c_val.upper()
        if c_upper.startswith('TOTAL') and ('CASES' in c_upper or 'CASE' in c_upper):
            break

        # "to be continued" → 스킵
        if 'to be continued' in c_val.lower():
            skip_until_data = True
            continue

        if skip_until_data:
            if 'DESCRIPTION' in c_val.upper() or 'QUANTITY' in c_val.upper():
                skip_until_data = False
            # 케이스 헤더 발견 시에도 스킵 해제
            if _parse_case_header(c_val):
                skip_until_data = False
            else:
                continue

        # 케이스 헤더 감지
        case_info = _parse_case_header(c_val)
        if case_info:
            case_no, total_cases = case_info

            # 이전 케이스 저장
            if current_case:
                cases.append(current_case)

            # 케이스 정보 파싱
            dims = _parse_dimensions(h_val)
            order_num = a_val if re.match(r'^\d{8}', a_val) else ''

            current_case = PackedCase(
                case_no=case_no,
                case_type="WOODEN CASE",
                items=[],
                dimensions=dims if dims else (0, 0, 0),
                case_weight=0,
                order_number=order_num,
            )

            # 케이스 헤더 행의 중량 정보
            if f_val:
                current_case._header_net_weight = f_val
            if g_val:
                current_case._header_gross_weight = g_val
                current_case.case_weight = g_val - (f_val or 0)

            continue

        # 카테고리 헤더
        if is_category_header(c_val) and e_val is None:
            current_category = c_val
            # CBM이 이 행에 있을 수 있음
            if h_val and not _parse_dimensions(h_val):
                try:
                    cbm_val = float(h_val.replace(',', ''))
                    if current_case and cbm_val < 100:
                        current_case._cbm = cbm_val
                except (ValueError, TypeError):
                    pass
            continue

        # 데이터 행: E열(수량)에 값이 있는 행
        if e_val is not None and e_val > 0 and c_val and current_case is not None:
            model = extract_model_from_description(c_val)
            ci_item = CILineItem(
                line_no=0,
                description=c_val,
                model_number=model or '',
                quantity=int(e_val),
                unit='SET',
                unit_price=0,
                amount=0,
                category_header=current_category,
            )
            net_w = f_val if f_val else 0
            packed_item = PackedItem(
                ci_item=ci_item,
                quantity=int(e_val),
                net_weight=net_w,
            )
            current_case.items.append(packed_item)

            # 첫 데이터 행의 중량/규격 정보 (케이스 전체)
            if len(current_case.items) == 1:
                dims = _parse_dimensions(h_val)
                if dims and current_case.dimensions == (0, 0, 0):
                    current_case.dimensions = dims
                if g_val and not hasattr(current_case, '_header_gross_weight'):
                    current_case._header_gross_weight = g_val
                    current_case.case_weight = (g_val or 0) - (f_val or 0)

            # 두 번째 행: CBM
            if len(current_case.items) == 2 and h_val:
                if not _parse_dimensions(h_val):
                    try:
                        cbm_val = float(h_val.replace(',', ''))
                        if cbm_val < 100:
                            current_case._cbm = cbm_val
                    except (ValueError, TypeError):
                        pass

    # 마지막 케이스 저장
    if current_case:
        cases.append(current_case)

    doc.cases = cases
    return doc
