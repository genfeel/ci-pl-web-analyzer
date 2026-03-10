"""CI(Commercial Invoice) 엑셀 파서

5개 샘플 분석 결과 기반:
- CI 컬럼: A=marks, C=description, F=quantity, G=unit_price, H=amount
- 헤더 위치가 파일마다 다름 → 동적 탐색
- 멀티 페이지: "- to be continued -" 이후 헤더 반복
- 합적 주문: S-TOTAL / G-TOTAL 패턴
"""
import re
from pathlib import Path
from typing import Optional

import openpyxl
import xlrd

from src.models.data_models import CILineItem, CIDocument, ProductCategory
from src.parser.model_number_parser import (
    extract_model_from_description,
    is_category_header,
    detect_category_from_header,
)


def _find_ci_sheet(sheet_names: list) -> Optional[str]:
    """CI 시트명 자동 탐지"""
    for name in sheet_names:
        upper = name.upper().strip()
        if upper == 'CI':
            return name
        if 'CI' in upper and 'PL' not in upper:
            return name
    for name in sheet_names:
        if 'CI' in name.upper():
            return name
    return sheet_names[0] if sheet_names else None


def _find_pl_sheet(sheet_names: list) -> Optional[str]:
    """PL 시트명 자동 탐지"""
    for name in sheet_names:
        upper = name.upper().strip()
        if upper == 'PL' or upper == 'P/L':
            return name
        if 'PL' in upper and 'CI' not in upper:
            return name
    for name in sheet_names:
        if 'PL' in name.upper() or 'P/L' in name.upper():
            return name
    return sheet_names[1] if len(sheet_names) > 1 else None


def _read_xlsx(filepath: str, sheet_name: str) -> list:
    """openpyxl로 .xlsx 읽기 → 2D 리스트 반환"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10, values_only=False):
        row_data = []
        for cell in row:
            row_data.append({
                'value': cell.value,
                'bold': cell.font.bold if cell.font else False,
            })
        rows.append(row_data)
    wb.close()
    return rows


def _read_xls(filepath: str, sheet_name: str) -> list:
    """xlrd로 .xls 읽기 → 2D 리스트 반환"""
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    ws = wb.sheet_by_name(sheet_name)
    rows = []
    for r in range(ws.nrows):
        row_data = []
        ncols = min(ws.ncols, 10)
        for c in range(ncols):
            cell = ws.cell(r, c)
            xf_index = ws.cell_xf_index(r, c)
            xf = wb.xf_list[xf_index]
            font = wb.font_list[xf.font_index]
            row_data.append({
                'value': cell.value if cell.value != '' else None,
                'bold': font.bold,
            })
        # 10열 미만이면 패딩
        while len(row_data) < 10:
            row_data.append({'value': None, 'bold': False})
        rows.append(row_data)
    wb.release_resources()
    return rows


def _cell_str(cell_dict: dict) -> str:
    """셀 값을 문자열로 변환"""
    val = cell_dict.get('value')
    if val is None:
        return ''
    return str(val).strip()


def _cell_num(cell_dict: dict) -> Optional[float]:
    """셀 값을 숫자로 변환"""
    val = cell_dict.get('value')
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace(',', '').strip()
        if cleaned:
            return float(cleaned)
    except (ValueError, TypeError):
        pass
    return None


def _find_header_row(rows: list) -> int:
    """컬럼 헤더 행 찾기 (Description of Goods 또는 Quantity 텍스트 기반)"""
    for i, row in enumerate(rows):
        for cell in row:
            text = _cell_str(cell).upper()
            if 'DESCRIPTION OF GOODS' in text or 'DESCRIPTION' in text:
                return i
            if 'QUANTITY' in text and any(
                'MARKS' in _cell_str(r).upper() or 'NO.' in _cell_str(r).upper()
                for r in row
            ):
                return i
    return 20  # 기본값


def _find_data_start(rows: list, header_row: int) -> int:
    """데이터 시작 행 찾기 (헤더 이후 첫 주문번호/카테고리 헤더/데이터 행)"""
    for i in range(header_row + 1, min(len(rows), header_row + 20)):
        row = rows[i]
        c_val = _cell_str(row[2]) if len(row) > 2 else ''

        # 주문번호 헤더 (합적 주문)
        if re.match(r'^\d{8}(-\d+)?$', c_val.strip()):
            return i

        if is_category_header(c_val):
            return i
        # 데이터 행: F열(수량)에 숫자가 있는 행
        f_val = _cell_num(row[5]) if len(row) > 5 else None
        if f_val is not None and f_val > 0:
            return i
    return header_row + 5


def parse_ci(filepath: str) -> CIDocument:
    """CI 엑셀 파일 파싱 → CIDocument 반환"""
    filepath = str(filepath)
    ext = Path(filepath).suffix.lower()

    # 시트 목록 가져오기
    if ext == '.xls':
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        sheet_names = wb.sheet_names()
        wb.release_resources()
    else:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

    ci_sheet = _find_ci_sheet(sheet_names)
    if not ci_sheet:
        raise ValueError(f"CI 시트를 찾을 수 없습니다: {sheet_names}")

    # 데이터 읽기
    if ext == '.xls':
        rows = _read_xls(filepath, ci_sheet)
    else:
        rows = _read_xlsx(filepath, ci_sheet)

    # 헤더 위치 탐색
    header_row = _find_header_row(rows)
    data_start = _find_data_start(rows, header_row)

    doc = CIDocument(
        filename=Path(filepath).name,
        header_info={
            'ci_sheet': ci_sheet,
            'pl_sheet': _find_pl_sheet(sheet_names),
            'header_row': header_row,
            'data_start': data_start,
        }
    )

    # 데이터 파싱
    items = []
    current_category = ''
    current_category_enum = ProductCategory.UNKNOWN
    current_order = ''
    order_numbers = []
    is_combined = False
    line_no = 0
    skip_until_next_page = False

    for i in range(data_start, len(rows)):
        row = rows[i]
        if len(row) < 8:
            continue

        a_val = _cell_str(row[0])
        c_val = _cell_str(row[2])
        d_val = _cell_str(row[3]) if len(row) > 3 else ''
        f_val = _cell_num(row[5])
        g_val = _cell_num(row[6])
        h_val = _cell_num(row[7])

        # "- to be continued -" 패턴: 페이지 구분 → 이후 헤더 스킵
        if 'to be continued' in c_val.lower() or 'to be continued' in d_val.lower():
            skip_until_next_page = True
            continue

        # 페이지 헤더 스킵 (COMMERCIAL INVOICE 타이틀 등)
        if skip_until_next_page:
            if 'DESCRIPTION' in c_val.upper() or 'QUANTITY' in _cell_str(row[5]).upper() if len(row) > 5 else False:
                skip_until_next_page = False
            continue

        # TOTAL / S-TOTAL / G-TOTAL 패턴
        c_upper = c_val.upper()
        if c_upper.startswith('TOTAL') or c_upper.startswith('G - TOTAL') or c_upper.startswith('G-TOTAL'):
            break  # 최종 합계 → 파싱 종료

        if c_upper.startswith('S - TOTAL') or c_upper.startswith('S-TOTAL') or c_upper.startswith('SUB-TOTAL'):
            is_combined = True
            continue  # 소계 행 스킵

        # 합적 주문번호 헤더 감지
        if _is_order_number_header(c_val, row):
            current_order = c_val.strip()
            if current_order and current_order not in order_numbers:
                order_numbers.append(current_order)
            is_combined = True
            continue

        # 카테고리 헤더 감지
        if is_category_header(c_val) and f_val is None:
            cat_name = detect_category_from_header(c_val)
            current_category = c_val.strip()
            current_category_enum = _map_category(cat_name)
            continue

        # 데이터 행: F열(수량)에 값이 있는 행
        if f_val is not None and f_val > 0 and c_val:
            line_no += 1
            model = extract_model_from_description(c_val)

            # 카테고리 세분화 (ACB → ACB_HGS / ACB_LARGE)
            category = current_category_enum
            if category == ProductCategory.ACB_HGS and model:
                if any(model.upper().startswith(p) for p in ('HGN', 'UAN', 'UCB')):
                    category = ProductCategory.ACB_LARGE

            item = CILineItem(
                line_no=line_no,
                description=c_val,
                model_number=model or c_val.split()[0] if c_val else '',
                quantity=int(f_val),
                unit=_cell_str(row[6]) if not isinstance(g_val, (int, float)) else 'SET',
                unit_price=g_val if g_val else 1.0,
                amount=h_val if h_val else f_val,
                category=category,
                category_header=current_category,
                parent_category=current_category if 'SPARE' in current_category.upper() else '',
                order_number=current_order,
            )
            items.append(item)

    doc.items = items
    doc.order_numbers = order_numbers if order_numbers else [doc.filename.split('-')[0]]
    doc.is_combined_order = is_combined

    return doc


def _is_order_number_header(text: str, row: list) -> bool:
    """합적 주문번호 헤더 행인지 판별

    패턴: 8자리 숫자 + 옵션 '-' + 숫자 (예: 25022297-1, 25021874)
    F열(수량)에 값이 없어야 함
    """
    text = text.strip()
    if not text:
        return False

    # 주문번호 패턴: 숫자 8자리 + 선택적 -숫자
    if re.match(r'^\d{8}(-\d+)?$', text):
        f_val = _cell_num(row[5]) if len(row) > 5 else None
        return f_val is None
    return False


def _map_category(cat_name: str) -> ProductCategory:
    """카테고리 문자열 → ProductCategory enum"""
    mapping = {
        'VCB': ProductCategory.VCB,
        'ACB': ProductCategory.ACB_HGS,  # 기본값, 이후 모델로 세분화
        'MCCB': ProductCategory.MCCB,
        'MC': ProductCategory.MC,
        'RELAY': ProductCategory.RELAY,
        'SPARE': ProductCategory.SPARE,
    }
    return mapping.get(cat_name, ProductCategory.UNKNOWN)


def get_sheet_names(filepath: str) -> list:
    """파일의 시트 목록 반환"""
    ext = Path(filepath).suffix.lower()
    if ext == '.xls':
        wb = xlrd.open_workbook(filepath)
        names = wb.sheet_names()
        wb.release_resources()
        return names
    else:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
