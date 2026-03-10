"""PL 생성기 테스트"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
import openpyxl

from src.models.data_models import (
    CILineItem, PackedCase, PackedItem, PLDocument, ProductCategory,
)
from src.generator.pl_generator import generate_pl_excel
from src.generator.format_rules import (
    format_dimensions, format_cbm, format_case_header, get_category_display_name,
)


class TestFormatRules:
    """서식 규칙 테스트"""

    def test_format_dimensions(self):
        assert format_dimensions((1100, 1100, 1120)) == "1100X1100X1120.MM"

    def test_format_dimensions_zero(self):
        assert format_dimensions((0, 0, 0)) == ""

    def test_format_cbm(self):
        assert format_cbm(1.3552) == "1.355"

    def test_format_case_header(self):
        assert format_case_header(1, 12) == "CASE NO : 1 of 12"

    def test_category_display_name_plural(self):
        name = get_category_display_name(ProductCategory.MCCB, plural=True)
        assert "BREAKERS" in name

    def test_category_display_name_singular(self):
        name = get_category_display_name(ProductCategory.MCCB, plural=False)
        assert "BREAKER" in name


class TestPLGenerator:
    """PL 생성기 테스트"""

    def _make_pl_doc(self):
        """테스트용 PLDocument 생성"""
        item = CILineItem(
            line_no=1, description="HGM100E-F3PTXXX",
            model_number="HGM100E", quantity=10, unit="SET",
            unit_price=1.0, amount=10.0,
            category=ProductCategory.MCCB,
            category_header="MOLDED CASE CIRCUIT BREAKER",
        )

        packed_item = PackedItem(ci_item=item, quantity=10, net_weight=18.0)

        case = PackedCase(
            case_no=1, case_type="PALLET",
            items=[packed_item],
            dimensions=(1100, 1100, 1120),
            case_weight=30.0,
            category=ProductCategory.MCCB,
            category_header="MOLDED CASE CIRCUIT BREAKER",
        )

        return PLDocument(
            filename="test_PL.xlsx",
            cases=[case],
            header_info={'invoice_no': 'TEST-001'},
            order_numbers=['TEST-001'],
        )

    def test_generate_creates_file(self):
        """파일 생성 확인"""
        pl_doc = self._make_pl_doc()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        result = generate_pl_excel(pl_doc, "", output_path)
        assert Path(result).exists()

        # 파일 읽기 확인
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        assert ws.title == "PL"
        assert ws.max_row > 0
        wb.close()

        Path(result).unlink()

    def test_generate_has_title(self):
        """PL 타이틀 확인"""
        pl_doc = self._make_pl_doc()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        generate_pl_excel(pl_doc, "", output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert "PACKING" in title.upper().replace(" ", "")
        wb.close()

        Path(output_path).unlink()

    def test_generate_has_total(self):
        """TOTAL 행 존재 확인"""
        pl_doc = self._make_pl_doc()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        generate_pl_excel(pl_doc, "", output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        found_total = False
        for row in ws.iter_rows(min_col=3, max_col=3, values_only=True):
            if row[0] and 'TOTAL' in str(row[0]).upper():
                found_total = True
                break
        assert found_total, "TOTAL 행을 찾을 수 없습니다"
        wb.close()

        Path(output_path).unlink()
